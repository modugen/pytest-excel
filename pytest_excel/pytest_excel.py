import re
from collections import OrderedDict
from datetime import datetime
from typing import Optional

import pandas as pd
import pytest
from _pytest.mark.structures import Mark
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

_py_ext_re = re.compile(r"\.py$")


def pytest_addoption(parser):
    group = parser.getgroup("terminal reporting")
    group.addoption(
        "--excelreport",
        "--excel-report",
        action="store",
        dest="excelpath",
        metavar="path",
        default=None,
        help="create excel report file at given path.",
    )


def pytest_configure(config):
    excelpath = config.option.excelpath
    if excelpath:
        config._excel = ExcelReporter(excelpath)
        config.pluginmanager.register(config._excel)


def pytest_unconfigure(config):
    excel = getattr(config, "_excel", None)
    if excel:
        del config._excel
        config.pluginmanager.unregister(excel)


def mangle_test_address(address):
    path, possible_open_bracket, params = address.partition("[")
    names = path.split("::")
    try:
        names.remove("()")
    except ValueError:
        pass

    names[0] = names[0].replace("/", ".")
    names[0] = _py_ext_re.sub("", names[0])
    names[-1] += possible_open_bracket + params
    return names


class ExcelReporter(object):
    STATUS_RESULT_MAP = {
        "PASSED": "OK",
        "FAILED": "ERR",
        "XPASSED": "OK",
        "XFAILED": "ERR",
        "SKIPPED": "ERR",
    }
    RESULT_COLOR_MAP = {
        "OK": "0099CC00",
        "ERR": "00FF0000",
    }
    """ Column and row index at which the result content of the sheet starts. """
    RESULT_START_IDX = 3
    SUMMARY_IDX = 2

    def __init__(self, excelpath):
        self.results = []
        self.result_matrix: Optional[pd.DataFrame] = None
        self.row_summaries: Optional[pd.DataFrame] = None
        self.col_summaries: Optional[pd.DataFrame] = None
        self.row_key = "model"
        self.column_key = "test_step"
        self.wbook = Workbook()
        self.wsheet = None
        self.rc = 1
        self.excelpath = datetime.now().strftime(excelpath)

    def append(self, result):
        self.results.append(result)

    def create_sheet(self):

        # TODO: check whether we really need to create a new sheet,
        #   the result.xls contains two sheets which is one too many
        self.wsheet = self.wbook.create_sheet(index=0)

        all_row_fields = self.result_matrix.index.insert(0, "summary")
        for i, row_label in enumerate(all_row_fields, self.SUMMARY_IDX):
            self.wsheet.cell(row=i, column=1).value = row_label

        all_col_fields = self.result_matrix.columns.insert(0, "summary")
        for i, col_label in enumerate(all_col_fields, self.SUMMARY_IDX):
            self.wsheet.cell(row=1, column=i).value = col_label

    def update_worksheet(self):
        self.update_results_in_worksheet()
        self.update_summaries_in_worksheet()

    def update_summaries_in_worksheet(self):
        assert (self.result_matrix.index == self.row_summaries.index).all()
        for row_idx, row_summary in enumerate(
                self.row_summaries.tolist(), self.RESULT_START_IDX
        ):
            cell = self.wsheet.cell(row=row_idx, column=self.SUMMARY_IDX)
            cell.value = row_summary
        for col_idx, col_summary in enumerate(
                self.col_summaries.tolist(), self.RESULT_START_IDX
        ):
            cell = self.wsheet.cell(row=self.SUMMARY_IDX, column=col_idx)
            cell.value = col_summary

    def update_results_in_worksheet(self):
        rows = [row for _, row in self.result_matrix.iterrows()]
        for row_index, row in enumerate(rows, self.RESULT_START_IDX):
            for col_index, value in enumerate(row.tolist(), self.RESULT_START_IDX):
                try:
                    cell = self.wsheet.cell(row=row_index, column=col_index)
                    cell.value = value
                    self.style_cell(cell, value)
                except ValueError:
                    pass

    def style_cell(self, cell, value):
        # green for "OK", red for "ERR"
        color_code = self.RESULT_COLOR_MAP[self.STATUS_RESULT_MAP[value]]
        cell.fill = PatternFill("solid", fgColor=color_code)
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def create_matrix(self):
        """Create a matrix filled with "MISSING" based on the row and column keys in self.results."""
        if self.result_matrix is not None:
            return

        row_headers = set([data[self.row_key] for data in self.results])
        col_headers = sorted(set([data[self.column_key] for data in self.results]))
        self.result_matrix = pd.DataFrame(
            "MISSING", index=row_headers, columns=col_headers
        )

    def fill_matrix(self):
        """Update the matrix with the data from self.results."""
        for data in self.results:
            value = data["result"]
            self.result_matrix.at[data[self.row_key], data[self.column_key]] = value

    def create_summaries(self):
        """Calculate row and column summaries based on self.result_matrix."""
        ok_or_err = self.result_matrix.applymap(lambda x: self.STATUS_RESULT_MAP[x])
        is_err = ok_or_err[ok_or_err == "ERR"]
        self.row_summaries = is_err.count(axis=1)
        self.col_summaries = is_err.count(axis=0)

    def sort_matrix(self):
        sorted_index = self.row_summaries.sort_values(ascending=False).index
        self.result_matrix = self.result_matrix.reindex(index=sorted_index)
        self.row_summaries = self.row_summaries.reindex(index=sorted_index)

    def save_excel(self):
        self.wbook.save(filename=self.excelpath)

    def build_result(self, report, status, message):

        result = OrderedDict()
        names = mangle_test_address(report.nodeid)

        model_name = None
        for mark in report.test_marker:
            if mark.name == "model":
                model_name = mark.args[0]
                break
        if model_name is None:
            raise AttributeError(
                f"Test {report.nodeid} is not marked with a model name.\n"
                "Add @pytest.mark.model('my_model_name') to the test function."
            )
        result["model"] = model_name

        # test name is something like 'test_excel_report_01[fzkhaus]'
        test_name = names[-1]
        step_name_parts = test_name.split("[")[0].split("_")[1:]
        result["test_step"] = "_".join(step_name_parts)

        if report.test_doc is None:
            result["description"] = report.test_doc
        else:
            result["description"] = report.test_doc.strip()

        result["result"] = status
        result["duration"] = getattr(report, "duration", 0.0)
        result["timestamp"] = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        result["message"] = message
        result["file_name"] = report.location[0]

        self.append(result)

    def append_pass(self, report):
        status = "PASSED"
        message = None
        self.build_result(report, status, message)

    def append_failure(self, report):
        if report.was_xfail:
            status = "XPASSED"
            message = "xfail-marked test passes"
        else:
            if hasattr(report.longrepr, "reprcrash"):
                message = report.longrepr.reprcrash.message
            elif isinstance(report.longrepr, str):
                message = report.longrepr
            else:
                message = str(report.longrepr)

            status = "FAILED"

        self.build_result(report, status, message)

    def append_error(self, report):

        message = report.longrepr
        status = "ERROR"
        self.build_result(report, status, message)

    def append_skipped(self, report):
        if report.was_xfail:
            status = "XFAILED"
            message = "expected test failure"
        else:
            status = "SKIPPED"
            _, _, message = report.longrepr
            if message.startswith("Skipped: "):
                message = message[9:]

        self.build_result(report, status, message)

    def build_tests(self, item):

        result = OrderedDict()
        names = mangle_test_address(item.nodeid)

        result["suite_name"] = names[-2]
        result["test_name"] = names[-1]
        if item.obj.__doc__ is None:
            result["description"] = item.obj.__doc__
        else:
            result["description"] = item.obj.__doc__.strip()
        result["file_name"] = item.location[0]
        test_marker = []
        test_message = []
        for k, v in item.keywords.items():
            if isinstance(v, list):
                for x in v:
                    if isinstance(x, Mark):
                        if x.name != "usefixtures":
                            test_marker.append(x.name)
                        if x.kwargs:
                            test_message.append(x.kwargs.get("reason"))

        test_markers = ", ".join(test_marker)
        result["markers"] = test_markers

        test_messages = ", ".join(test_message)
        result["message"] = test_messages
        self.append(result)

    def append_tests(self, item):

        self.build_tests(item)

    @pytest.mark.trylast
    def pytest_collection_modifyitems(self, session, config, items):
        """called after collection has been performed, may filter or re-order
        the items in-place."""
        if session.config.option.collectonly:
            for item in items:
                self.append_tests(item)

    @pytest.mark.hookwrapper
    def pytest_runtest_makereport(self, item, call):

        outcome = yield

        report = outcome.get_result()
        report.test_doc = item.obj.__doc__

        test_marker = []
        for k, v in item.keywords.items():
            if isinstance(v, list):
                for x in v:
                    if isinstance(x, Mark):
                        test_marker.append(x)
            if isinstance(v, Mark):
                test_marker.append(v)
        report.test_marker = test_marker

        xfail_markers = [mark for mark in item.own_markers if mark.name == "xfail"]
        report.was_xfail = True if xfail_markers else False

    def pytest_runtest_logreport(self, report):

        if report.passed:
            if report.when == "call":  # ignore setup/teardown
                self.append_pass(report)

        elif report.failed:
            if report.when == "call":
                self.append_failure(report)

            else:
                self.append_error(report)

        elif report.skipped:
            self.append_skipped(report)

    def pytest_sessionfinish(self, session):
        if not hasattr(session.config, "slaveinput"):
            if self.results:
                self.create_matrix()
                self.fill_matrix()
                self.create_summaries()
                self.sort_matrix()

                self.create_sheet()
                self.update_worksheet()
                self.save_excel()

    def pytest_terminal_summary(self, terminalreporter):
        if self.results:
            terminalreporter.write_sep("-", "excel report: %s" % self.excelpath)
